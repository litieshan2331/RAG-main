from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from sqlalchemy import select, text
from sqlalchemy.orm import Session

from app.models.entities import TableMeta
from app.storage.minio_store import safe_object_part


TABLE_SUFFIXES = {".csv", ".xlsx", ".xls"}


@dataclass(frozen=True)
class ColumnSpec:
    name: str
    original_name: str
    sql_type: str


@dataclass(frozen=True)
class TableData:
    table_name: str
    columns: list[ColumnSpec]
    rows: list[dict[str, Any]]
    create_sql: str
    columns_info: str


class StructuredTableIngestor:
    def __init__(self, db: Session) -> None:
        self.db = db

    def ingest(self, *, file_name: str, data: bytes, table_name: str | None, doc_id: int) -> TableData:
        rows = parse_table_file(file_name, data)
        if not rows:
            raise RuntimeError("structured table contains no rows")

        resolved_table_name = normalize_table_name(table_name or f"structured_{doc_id}_{Path(file_name).stem}")
        columns = infer_columns(rows)
        create_sql = build_create_sql(resolved_table_name, columns)
        normalized_rows = normalize_rows(rows, columns, doc_id)

        self.db.execute(text(create_sql))
        if normalized_rows:
            insert_sql = build_insert_sql(resolved_table_name, ["source_doc_id", "row_number", *[column.name for column in columns]])
            self.db.execute(text(insert_sql), normalized_rows)

        columns_info = json.dumps(
            [
                {
                    "name": column.name,
                    "originalName": column.original_name,
                    "type": column.sql_type,
                }
                for column in columns
            ],
            ensure_ascii=False,
        )
        self._upsert_table_meta(
            resolved_table_name,
            description=f"Structured table imported from {file_name}",
            create_sql=create_sql,
            columns_info=columns_info,
        )
        self.db.commit()
        return TableData(
            table_name=resolved_table_name,
            columns=columns,
            rows=normalized_rows,
            create_sql=create_sql,
            columns_info=columns_info,
        )

    def _upsert_table_meta(self, table_name: str, description: str, create_sql: str, columns_info: str) -> None:
        table_meta = self.db.execute(
            select(TableMeta).where(TableMeta.table_name == table_name, TableMeta.deleted == 0)
        ).scalar_one_or_none()
        if table_meta is None:
            self.db.add(
                TableMeta(
                    table_name=table_name,
                    description=description,
                    create_sql=create_sql,
                    columns_info=columns_info,
                )
            )
            return

        table_meta.description = description
        table_meta.create_sql = create_sql
        table_meta.columns_info = columns_info


def parse_table_file(file_name: str, data: bytes) -> list[dict[str, Any]]:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        return parse_csv(data)
    if suffix in {".xlsx", ".xls"}:
        return parse_excel(data)
    raise RuntimeError(f"unsupported structured table file type: {suffix}")


def parse_csv(data: bytes) -> list[dict[str, Any]]:
    text_value = decode_text(data)
    sample = text_value[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text_value), dialect=dialect)
    return [{key or "column": value for key, value in row.items()} for row in reader]


def parse_excel(data: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Excel ingestion requires openpyxl. Run `uv sync` first.") from exc

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_index = next((index for index, row in enumerate(rows) if any(cell is not None for cell in row)), None)
    if header_index is None:
        return []

    headers = [str(value).strip() if value is not None and str(value).strip() else f"column_{index + 1}" for index, value in enumerate(rows[header_index])]
    output: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        output.append({headers[index]: value for index, value in enumerate(row[: len(headers)])})
    return output


def decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def normalize_table_name(value: str) -> str:
    cleaned = safe_object_part(value).lower()
    cleaned = re.sub(r"[^a-z0-9_]+", "_", cleaned)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    if not cleaned or not cleaned[0].isalpha():
        cleaned = f"t_{cleaned or 'table'}"
    return cleaned[:64]


def normalize_column_name(value: str, index: int, used: set[str]) -> str:
    cleaned = value.strip().lower()
    cleaned = re.sub(r"[^a-z0-9_]+", "_", cleaned)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    if not cleaned or not cleaned[0].isalpha():
        cleaned = f"col_{index + 1}"

    candidate = cleaned[:58]
    suffix = 2
    while candidate in used:
        candidate = f"{cleaned[:54]}_{suffix}"
        suffix += 1
    used.add(candidate)
    return candidate


def infer_columns(rows: list[dict[str, Any]]) -> list[ColumnSpec]:
    headers = list(rows[0].keys())
    used: set[str] = set()
    columns: list[ColumnSpec] = []
    for index, header in enumerate(headers):
        values = [row.get(header) for row in rows]
        columns.append(
            ColumnSpec(
                name=normalize_column_name(str(header), index, used),
                original_name=str(header),
                sql_type=infer_sql_type(values),
            )
        )
    return columns


def infer_sql_type(values: list[Any]) -> str:
    non_empty = [value for value in values if value is not None and str(value).strip() != ""]
    if not non_empty:
        return "TEXT"
    if all(isinstance(value, bool) or str(value).strip().lower() in {"true", "false"} for value in non_empty):
        return "TINYINT"
    if all(is_int(value) for value in non_empty):
        return "BIGINT"
    if all(is_decimal(value) for value in non_empty):
        return "DOUBLE"
    if all(isinstance(value, date | datetime) for value in non_empty):
        return "DATETIME"
    max_len = max(len(str(value)) for value in non_empty)
    if max_len <= 512:
        return "VARCHAR(512)"
    return "TEXT"


def is_int(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    try:
        int(str(value).strip())
        return "." not in str(value).strip()
    except ValueError:
        return False


def is_decimal(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    try:
        Decimal(str(value).strip())
        return True
    except (InvalidOperation, ValueError):
        return False


def normalize_rows(rows: list[dict[str, Any]], columns: list[ColumnSpec], doc_id: int) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows, start=1):
        item = {"source_doc_id": doc_id, "row_number": row_number}
        for column in columns:
            value = row.get(column.original_name)
            item[column.name] = coerce_value(value, column.sql_type)
        normalized.append(item)
    return normalized


def coerce_value(value: Any, sql_type: str) -> Any:
    if value is None:
        return None
    if isinstance(value, str) and value.strip() == "":
        return None
    if sql_type == "TINYINT":
        if isinstance(value, bool):
            return int(value)
        return 1 if str(value).strip().lower() == "true" else 0
    if sql_type == "BIGINT":
        return int(str(value).strip())
    if sql_type == "DOUBLE":
        return float(str(value).strip())
    if sql_type == "DATETIME" and isinstance(value, date | datetime):
        return value
    return str(value)


def build_create_sql(table_name: str, columns: list[ColumnSpec]) -> str:
    column_defs = [
        "`id` BIGINT NOT NULL AUTO_INCREMENT",
        "`source_doc_id` BIGINT NOT NULL",
        "`row_number` INT NOT NULL",
        *[f"`{column.name}` {column.sql_type} NULL" for column in columns],
        "`created_at` DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP",
        "PRIMARY KEY (`id`)",
        "INDEX `idx_source_doc_id` (`source_doc_id`)",
    ]
    return f"CREATE TABLE IF NOT EXISTS `{table_name}` (\n    " + ",\n    ".join(column_defs) + "\n) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci"


def build_insert_sql(table_name: str, columns: list[str]) -> str:
    identifiers = ", ".join(f"`{column}`" for column in columns)
    placeholders = ", ".join(f":{column}" for column in columns)
    return f"INSERT INTO `{table_name}` ({identifiers}) VALUES ({placeholders})"
